"""
Excel Formula Recalculation Script
Recalculates all formulas in an Excel file using LibreOffice

支持平台：Windows, macOS, Linux
"""

import json
import os
import platform
import subprocess
import sys
from pathlib import Path

from office.soffice import get_soffice_env, find_soffice_path

from openpyxl import load_workbook

# LibreOffice macro 存储路径（按平台区分）
MACRO_DIR_WINDOWS = os.path.join(
    os.environ.get("APPDATA", ""),
    "LibreOffice", "4", "user", "basic", "Standard"
)
MACRO_DIR_MACOS = "~/Library/Application Support/LibreOffice/4/user/basic/Standard"
MACRO_DIR_LINUX = "~/.config/libreoffice/4/user/basic/Standard"
MACRO_FILENAME = "Module1.xba"

RECALCULATE_MACRO = """<?xml version="1.0" encoding="UTF-8"?>
<!DOCTYPE script:module PUBLIC "-//OpenOffice.org//DTD OfficeDocument 1.0//EN" "module.dtd">
<script:module xmlns:script="http://openoffice.org/2000/script" script:name="Module1" script:language="StarBasic">
    Sub RecalculateAndSave()
      ThisComponent.calculateAll()
      ThisComponent.store()
      ThisComponent.close(True)
    End Sub
</script:module>"""


def _get_macro_dir() -> str:
    """获取当前平台的 LibreOffice macro 存储目录"""
    system = platform.system()
    if system == "Windows":
        return MACRO_DIR_WINDOWS
    elif system == "Darwin":
        return os.path.expanduser(MACRO_DIR_MACOS)
    else:
        return os.path.expanduser(MACRO_DIR_LINUX)


def setup_libreoffice_macro():
    """
    安装 LibreOffice RecalculateAndSave 宏

    首次运行时自动创建 macro 目录并写入重新计算宏。
    跨平台支持 Windows、macOS 和 Linux。
    """
    macro_dir = _get_macro_dir()
    macro_file = os.path.join(macro_dir, MACRO_FILENAME)

    if (
        os.path.exists(macro_file)
        and "RecalculateAndSave" in Path(macro_file).read_text()
    ):
        return True

    if not os.path.exists(macro_dir):
        # 需要先启动一次 LibreOffice 以初始化用户配置目录
        soffice = find_soffice_path()
        try:
            subprocess.run(
                [soffice, "--headless", "--terminate_after_init"],
                capture_output=True,
                timeout=15,
                env=get_soffice_env(),
            )
        except subprocess.TimeoutExpired:
            pass  # LibreOffice 初始化可能超时，但目录通常已创建
        os.makedirs(macro_dir, exist_ok=True)

    try:
        Path(macro_file).write_text(RECALCULATE_MACRO)
        return True
    except Exception:
        return False


def recalc(filename, timeout=30):
    """
    使用 LibreOffice 重新计算 Excel 文件中的所有公式

    跨平台实现：
    - Windows: 使用 Python subprocess.run 的 timeout 参数
    - Linux: 使用系统 timeout 命令
    - macOS: 使用 gtimeout（如果可用）

    Args:
        filename: Excel 文件路径
        timeout: 超时时间（秒）

    Returns:
        包含状态、错误数量和公式数量的字典
    """
    if not Path(filename).exists():
        return {"error": f"File {filename} does not exist"}

    abs_path = str(Path(filename).absolute())

    if not setup_libreoffice_macro():
        return {"error": "Failed to setup LibreOffice macro"}

    soffice = find_soffice_path()

    cmd = [
        soffice,
        "--headless",
        "--norestore",
        "vnd.sun.star.script:Standard.Module1.RecalculateAndSave?language=Basic&location=application",
        abs_path,
    ]

    system = platform.system()

    if system == "Windows":
        # Windows 上使用 Python 内置的 subprocess timeout 机制
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                env=get_soffice_env(),
                timeout=timeout,
            )
        except subprocess.TimeoutExpired:
            return {"error": f"LibreOffice recalculation timed out after {timeout}s"}
    else:
        # Linux/macOS 使用系统级 timeout 命令
        if system == "Linux":
            cmd = ["timeout", str(timeout)] + cmd
        elif system == "Darwin" and _has_gtimeout():
            cmd = ["gtimeout", str(timeout)] + cmd

        result = subprocess.run(cmd, capture_output=True, text=True, env=get_soffice_env())

    if result.returncode != 0 and result.returncode != 124:
        error_msg = result.stderr or "Unknown error during recalculation"
        if "Module1" in error_msg or "RecalculateAndSave" not in error_msg:
            return {"error": "LibreOffice macro not configured properly"}
        return {"error": error_msg}

    try:
        wb = load_workbook(filename, data_only=True)

        excel_errors = [
            "#VALUE!",
            "#DIV/0!",
            "#REF!",
            "#NAME?",
            "#NULL!",
            "#NUM!",
            "#N/A",
        ]
        error_details = {err: [] for err in excel_errors}
        total_errors = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None and isinstance(cell.value, str):
                        for err in excel_errors:
                            if err in cell.value:
                                location = f"{sheet_name}!{cell.coordinate}"
                                error_details[err].append(location)
                                total_errors += 1
                                break

        wb.close()

        result = {
            "status": "success" if total_errors == 0 else "errors_found",
            "total_errors": total_errors,
            "error_summary": {},
        }

        for err_type, locations in error_details.items():
            if locations:
                result["error_summary"][err_type] = {
                    "count": len(locations),
                    "locations": locations[:20],
                }

        wb_formulas = load_workbook(filename, data_only=False)
        formula_count = 0
        for sheet_name in wb_formulas.sheetnames:
            ws = wb_formulas[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if (
                        cell.value
                        and isinstance(cell.value, str)
                        and cell.value.startswith("=")
                    ):
                        formula_count += 1
        wb_formulas.close()

        result["total_formulas"] = formula_count

        return result

    except Exception as e:
        return {"error": str(e)}


def _has_gtimeout():
    """检查 macOS 上是否安装了 gtimeout（来自 coreutils）"""
    try:
        subprocess.run(
            ["gtimeout", "--version"], capture_output=True, timeout=1, check=False
        )
        return True
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False


def main():
    if len(sys.argv) < 2:
        print("Usage: python recalc.py <excel_file> [timeout_seconds]")
        print("\nRecalculates all formulas in an Excel file using LibreOffice")
        print("\nReturns JSON with error details:")
        print("  - status: 'success' or 'errors_found'")
        print("  - total_errors: Total number of Excel errors found")
        print("  - total_formulas: Number of formulas in the file")
        print("  - error_summary: Breakdown by error type with locations")
        print("    - #VALUE!, #DIV/0!, #REF!, #NAME?, #NULL!, #NUM!, #N/A")
        sys.exit(1)

    filename = sys.argv[1]
    timeout = int(sys.argv[2]) if len(sys.argv) > 2 else 30

    result = recalc(filename, timeout)
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
